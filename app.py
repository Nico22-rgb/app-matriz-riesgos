import streamlit as st
import pandas as pd
import io
import base64
from PIL import Image
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

# Configuración de página
st.set_page_config(page_title="Análisis de Riesgos", layout="centered")

st.markdown("<h1 style='text-align: center;'>Análisis de Riesgos - Área de Validaciones</h1>", unsafe_allow_html=True)

def mostrar_logo_adaptable(path_png_transparente):
    try:
        with open(path_png_transparente, "rb") as image_file:
            encoded = image_file.read()
        st.markdown(
            f"""
            <div style="display: flex; justify-content: center;">
                <img src="data:image/png;base64,{base64.b64encode(encoded).decode()}" width="300">
            </div>
            """,
            unsafe_allow_html=True
        )
    except Exception as e:
        st.warning(f"No se pudo cargar el logo. Error: {e}")
        st.info("Verifica que el archivo exista y esté en formato PNG transparente.")

mostrar_logo_adaptable("altea.png")

@st.cache_data
def load_excel(file, sheet_name):
    wb = load_workbook(file)
    ws = wb[sheet_name]
    data = []
    headers = [cell.value for cell in ws[1] if cell.value]

    last_etapa = None
    for row in ws.iter_rows(min_row=2):
        row_data = []
        for cell_idx, cell in enumerate(row):
            if cell.value is not None:
                if headers[cell_idx] == "Etapa":
                    last_etapa = cell.value
                row_data.append(cell.value)
            else:
                if headers[cell_idx] == "Etapa" and last_etapa is not None:
                    row_data.append(last_etapa)
                else:
                    row_data.append(None)
        if any(x is not None for x in row_data):
            data.append(row_data[:len(headers)])
    df = pd.DataFrame(data, columns=headers)
    if "Etapa" in df.columns and "Operación" in df.columns and "Atributo" in df.columns:
        df = df[["Etapa", "Operación", "Atributo"]].copy()
    return df

CONTRASENA = "M"
if "autenticado" not in st.session_state:
    st.session_state.autenticado = False

if not st.session_state.autenticado:
    st.markdown("## 🔐 Ingreso restringido")
    contrasena = st.text_input("Ingrese la contraseña para continuar:", type="password")
    if contrasena == CONTRASENA:
        st.session_state.autenticado = True
        st.success("✅ Acceso concedido.")
    else:
        if contrasena != "":
            st.error("❌ Contraseña incorrecta.")
        st.stop()

st.markdown("<h5>Por favor sube el archivo de la base de datos de las matrices de riesgo</h5>", unsafe_allow_html=True)
archivo = st.file_uploader("", type=["xlsx"], key="file_uploader_unique")

if "ET_OP_AT_df" not in st.session_state and archivo is not None:
    try:
        ET_OP_AT_df = load_excel(archivo, sheet_name="ET_OP_AT")
        st.session_state['ET_OP_AT_df'] = ET_OP_AT_df
    except Exception as e:
        st.warning(f"No se pudo cargar la hoja 'ET_OP_AT'. Error: {e}. Usando valores predeterminados.")
        st.session_state['ET_OP_AT_df'] = pd.DataFrame({
            "Etapa": ["Verificación", "Pesaje", "Limpieza", "Mezcla"],
            "Operación": ["Prueba 1", "Prueba 2", "Inspección", "Mezcla"],
            "Atributo": ["Dimensión", "Peso", "Pureza", "Humedad"]
        })

if archivo:
    tipo_validacion = st.selectbox("Seleccione el tipo de validación a realizar", [
        "Validación de procesos", "Validación de campaña", "Validación de limpieza"
    ], index=None)

    etapas_seleccionadas = []
    sheet_to_use = None
    if tipo_validacion in ["Validación de procesos", "Validación de campaña"]:
        tipo_linea = st.selectbox("¿A qué línea de fabricación pertenece su producto?", [
            "Línea de medicamentos sólidos",
            "Línea de medicamentos líquidos y semisólidos"
        ], index=None)

        if tipo_linea == "Línea de medicamentos sólidos":
            sheet_to_use = "MR 1 sólidos"
            etapas_posibles = [
                "Verificación de prerrequisitos de validación", "Pesaje/Dispensación de materias primas",
                "Pulverización", "Pelletización", "Granulacion", "Secado", "Compactación",
                "Mezcla (Lubricación)", "Encapsulado", "Compresión", "Recubrimiento", "Grageado",
                "Revisión", "Envase blíster", "Envase foil", "Envase frasco", "Envase sobre",
                "Envase tubo", "Empaque blíster", "Empaque manual foil", "Empaque frasco",
                "Empaque tubo", "Recogida de blísters", "Codificado manual"
            ]
            for etapa in etapas_posibles:
                if st.toggle(etapa):
                    etapas_seleccionadas.append(etapa)
        elif tipo_linea == "Línea de medicamentos líquidos y semisólidos":
            sheet_to_use = "MR 1 líquidos y semisólidos"
            etapas_liquidas = [
                "Verificación de prerrequisitos de validación", "Pesaje/Dispensación de materias primas",
                "Disolución/Dispersión", "Homogenización", "Filtración", "Envase frascos",
                "Envase sobres", "Envase tubos", "Empaque manual frascos", "Empaque manual sobre",
                "Empaque tubos"
            ]
            for etapa in etapas_liquidas:
                if st.toggle(etapa):
                    etapas_seleccionadas.append(etapa)

    elif tipo_validacion == "Validación de limpieza":
        sheet_to_use = "MR 1 limpieza"
        etapas_limpieza = [
            "Verificación de prerrequisitos de validación",
            "Limpieza preliminar y desmonte del equipo (piezas móviles)",
            "Limpieza de piezas móviles y parte interna de los equipos",
            "Seguimiento al proceso de limpieza",
            "Uso, desmonte y prelavado de las mangas",
            "Verificación y limpieza de las mangas"
        ]
        for etapa in etapas_limpieza:
            if st.toggle(etapa):
                etapas_seleccionadas.append(etapa)

    if etapas_seleccionadas and sheet_to_use:
        if "excel_buffer" not in st.session_state:
            if st.button("Generar matriz de riesgo con Excel"):
                try:
                    df = load_excel(archivo, sheet_to_use)
                    rangos_por_hoja = {
                        "MR 1 sólidos": {etapa: (i, i+1) for i, etapa in enumerate([
                            "Verificación de prerrequisitos de validación", "Pesaje/Dispensación de materias primas",
                            "Pulverización", "Pelletización", "Granulacion", "Secado", "Compactación",
                            "Mezcla (Lubricación)", "Encapsulado", "Compresión", "Recubrimiento", "Grageado",
                            "Revisión", "Envase blíster", "Envase foil", "Envase frasco", "Envase sobre",
                            "Envase tubo", "Empaque blíster", "Empaque manual foil", "Empaque frasco",
                            "Empaque tubo", "Recogida de blísters", "Codificado manual"
                        ])},
                        "MR 1 líquidos y semisólidos": {etapa: (i, i+1) for i, etapa in enumerate([
                            "Verificación de prerrequisitos de validación", "Pesaje/Dispensación de materias primas",
                            "Disolución/Dispersión", "Homogenización", "Filtración", "Envase frascos",
                            "Envase sobres", "Envase tubos", "Empaque manual frascos", "Empaque manual sobre", "Empaque tubos"
                        ])},
                        "MR 1 limpieza": {etapa: (i, i+1) for i, etapa in enumerate([
                            "Verificación de prerrequisitos de validación",
                            "Limpieza preliminar y desmonte del equipo (piezas móviles)",
                            "Limpieza de piezas móviles y parte interna de los equipos",
                            "Seguimiento al proceso de limpieza",
                            "Uso, desmonte y prelavado de las mangas",
                            "Verificación y limpieza de las mangas"
                        ])}
                    }
                    rangos_para_hoja_actual = rangos_por_hoja.get(sheet_to_use, {})
                    if not rangos_para_hoja_actual:
                        st.error(f"No se encontraron rangos definidos para la hoja '{sheet_to_use}'.")
                        st.stop()
                    encabezado = df.iloc[[0]]
                    bloques = []
                    for etapa in etapas_seleccionadas:
                        if etapa in rangos_para_hoja_actual:
                            inicio, fin = rangos_para_hoja_actual[etapa]
                            bloques.append(df.iloc[inicio:fin + 1])
                        else:
                            st.warning(f"La etapa '{etapa}' no tiene rangos definidos.")
                    tabla = pd.concat([encabezado] + bloques, ignore_index=True)
                    buffer = io.BytesIO()
                    tabla.to_excel(buffer, index=False)
                    buffer.seek(0)

                    wb = load_workbook(buffer)
                    ws = wb.active

                    palegreen_fill = PatternFill(start_color="C0E080", end_color="C0E080", fill_type="solid")
                    bold_font = Font(bold=True)
                    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    for cell in ws[1]:
                        cell.fill = palegreen_fill
                        cell.font = bold_font
                        cell.alignment = center_alignment

                    columnas_a_combinar = [1, 3, 4]
                    for col_to_merge in columnas_a_combinar:
                        current_value = None
                        start_row = 2
                        for row in range(2, ws.max_row + 1):
                            value = ws.cell(row=row, column=col_to_merge).value
                            value = str(value).strip() if value is not None else ""
                            if row == 2:
                                current_value = value
                                continue
                            if value != current_value or row == ws.max_row:
                                if row - start_row > 1 or (row == ws.max_row and value == current_value):
                                    end_row = row if value != current_value else row + 1
                                    ws.merge_cells(
                                        start_row=start_row,
                                        start_column=col_to_merge,
                                        end_row=end_row - 1,
                                        end_column=col_to_merge
                                    )
                                    ws.cell(row=start_row, column=col_to_merge).alignment = Alignment(
                                        horizontal="center", vertical="center", wrap_text=True
                                    )
                                current_value = value
                                start_row = row

                    for r_idx in range(2, ws.max_row + 1):
                        ws[f"O{r_idx}"].value = f"=ROUND(POWER((J{r_idx}*L{r_idx}*N{r_idx}),1/3),1)"
                        ws[f"P{r_idx}"].value = f'=IF(O{r_idx}<1.33,"Bajo",IF(O{r_idx}<3,"Moderado","Alto"))'
                        ws[f"Q{r_idx}"].value = f'=IF(P{r_idx}="Alto","Alta",IF(P{r_idx}="Moderado","Media","Baja"))'

                    rojo = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    naranja = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

                    ws.conditional_formatting.add("P2:P" + str(ws.max_row), FormulaRule(formula=['P2="Alto"'], fill=rojo))
                    ws.conditional_formatting.add("P2:P" + str(ws.max_row), FormulaRule(formula=['P2="Moderado"'], fill=naranja))
                    ws.conditional_formatting.add("P2:P" + str(ws.max_row), FormulaRule(formula=['P2="Bajo"'], fill=verde))
                    ws.conditional_formatting.add("Q2:Q" + str(ws.max_row), FormulaRule(formula=['Q2="Alta"'], fill=rojo))
                    ws.conditional_formatting.add("Q2:Q" + str(ws.max_row), FormulaRule(formula=['Q2="Media"'], fill=naranja))
                    ws.conditional_formatting.add("Q2:Q" + str(ws.max_row), FormulaRule(formula=['Q2="Baja"'], fill=verde))

                    for column in ws.columns:
                        column_letter = get_column_letter(column[0].column)
                        if column_letter == 'P':
                            ws.column_dimensions[column_letter].width = 30
                        elif column_letter == 'Q':
                            ws.column_dimensions[column_letter].width = 14
                        else:
                            max_length = 0
                            for cell in column:
                                try:
                                    if cell.value is not None:
                                        cell_value_str = str(cell.value)
                                        max_length = max(max_length, len(cell_value_str))
                                except Exception:
                                    pass
                            ws.column_dimensions[column_letter].width = max_length + 3

                    for row in ws.iter_rows():
                        for cell in row:
                            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                    output = io.BytesIO()
                    wb.save(output)
                    output.seek(0)
                    st.session_state['excel_buffer'] = output
                    st.session_state['etapas_seleccionadas'] = etapas_seleccionadas

                    st.balloons()
                    st.success("¡Matriz de riesgo generada con éxito!")
                except Exception as e:
                    st.error(f"Ocurrió un error al procesar el archivo: {e}")

        if "excel_buffer" in st.session_state:
            st.download_button(
                label="Descargar matriz de riesgo 📥",
                data=st.session_state['excel_buffer'].getvalue(),
                file_name="matriz_riesgo.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                on_click=lambda: st.session_state.update({"descarga_realizada": True})
            )
            if st.session_state.get("descarga_realizada", False):
                st.markdown("<h5>De acuerdo con la matriz de riesgo generada, seleccione las etapas que involucran operaciones con criticidad alta:</h5>", unsafe_allow_html=True)
                selected_alta = st.multiselect(
                    "Seleccione las etapas que involucran operaciones con criticidad alta:",
                    options=st.session_state['etapas_seleccionadas'],
                    key="alta_seleccion"
                )
                if selected_alta:
                    operaciones_texto = ', '.join(selected_alta)
                    st.info(f"Las etapas que involucran operaciones críticas son: {operaciones_texto}")
                    if st.button("Generar matriz de priorización del riesgo"):
                        st.session_state['mostrar_matriz'] = True

if st.session_state.get('mostrar_matriz', False):
    def mostrar_imagen_con_zonas(path_png_transparente):
        try:
            with open(path_png_transparente, "rb") as image_file:
                encoded = base64.b64encode(image_file.read()).decode()
            st.markdown(
                f"""
                <div style="display: flex; justify-content: center;">
                    <img src="data:image/png;base64,{encoded}" width="800">
                </div>
                """,
                unsafe_allow_html=True
            )
        except Exception as e:
            st.error(f"Error al cargar la imagen: {e}")
    mostrar_imagen_con_zonas("Matriz de priorizacion de riesgos.png")
    st.markdown('**Ubica en la matriz el nivel de riesgo asociado a tu operación/etapa para saber si es necesario implementar controles adicionales durante la validación.**')

    if "area_roja_consultada" not in st.session_state:
        st.session_state['area_roja_consultada'] = False

    with st.expander("🟢 Área Verde"):
        st.write("No es necesario implementar controles adicionales en esta operación-etapa para demostrar que la verificación a ser efectuada es lo suficientemente robusta para dar un concepto final. Se recomienda monitoreo rutinario.")
    with st.expander("🟡 Área Amarilla"):
        st.write("Considera implementar acciones o controles adicionales durante los seguimientos de validación para demostrar que la verificación a ser efectuada es lo suficientemente robusta para dar un concepto final.")
    with st.expander("🔴 Área Roja"):
        st.write("Es necesario implementar acciones o controles adicionales durante los seguimientos de validación para garantizar que la verificación a ser efectuada es lo suficientemente robusta para dar un concepto final.")
        if not st.session_state.get('area_roja_consultada', False):
            st.session_state['area_roja_consultada'] = True
            st.rerun()

if st.session_state.get('area_roja_consultada', False):
    st.markdown("""
        <h3 style='text-align: center; color: #2c3e50; margin-bottom: 30px;'>Plan de Muestreo</h3>
    """, unsafe_allow_html=True)
    st.markdown(f"""
        <div style="display: flex; align-items: center; justify-content: center; gap: 30px; margin-bottom: 35px;">
            <img src="data:image/png;base64,{base64.b64encode(open('muestreo.png', 'rb').read()).decode()}"
                 style="width: 130px; height: auto; border-radius: 5px;" />
            <p style="text-align: justify; max-width: 600px;">
                La elaboración de esta propuesta de plan de muestreo está basada en la fórmula para el cálculo del tamaño de muestra
                para poblaciones de tamaño finito del capítulo <b>&lt;1010&gt;</b> de la Farmacopea de los Estados Unidos de América, y en
                las recomendaciones del procedimiento estándar de operación <b>PSO-VAL-007</b> sobre el muestreo y análisis de datos en la
                validación de procesos de manufactura en Altea Farmacéutica S.A.
            </p>
        </div>
    """, unsafe_allow_html=True)
    st.markdown("<div style='margin-bottom: 20px;'></div>", unsafe_allow_html=True)
    with st.expander("Por favor diligencia los campos correspondientes basado en la matriz de riesgo obtenida para el proceso ", expanded=True):
        ET_OP_AT_df = st.session_state.get('ET_OP_AT_df', pd.DataFrame())
        if ET_OP_AT_df.empty:
            st.warning("No se cargaron datos de la hoja 'ET_OP_AT'. Usando valores predeterminados.")
            etapas_disponibles = ["Verificación", "Pesaje", "Limpieza", "Mezcla"]
        else:
            etapas_disponibles = ET_OP_AT_df["Etapa"].dropna().unique().tolist()
        col1, col2, col3 = st.columns(3)
        with col1:
            etapa = st.selectbox("Etapa", options=etapas_disponibles)
        if not ET_OP_AT_df.empty:
            ET_OP_AT_df["Etapa_normalized"] = ET_OP_AT_df["Etapa"].str.strip().str.lower()
            etapa_normalizada = etapa.strip().lower()
            filas_filtradas = ET_OP_AT_df[ET_OP_AT_df["Etapa_normalized"] == etapa_normalizada]
            operaciones_filtradas = filas_filtradas["Operación"].dropna().tolist()
            atributos_filtrados = filas_filtradas["Atributo"].dropna().tolist()
        else:
            operaciones_filtradas = ["Prueba 1", "Prueba 2", "Inspección"]
            atributos_filtrados = ["Dimensión", "Peso", "Pureza"]
        if not operaciones_filtradas:
            st.warning(f"No se encontraron operaciones para la etapa '{etapa}'. Verifica la hoja 'ET_OP_AT'.")
            operaciones_filtradas = ["Sin opciones"]
        if not atributos_filtrados:
            st.warning(f"No se encontraron atributos para la etapa '{etapa}'. Verifica la hoja 'ET_OP_AT'.")
            atributos_filtrados = ["Sin opciones"]
        with col2:
            operacion = st.selectbox("Operación", options=operaciones_filtradas)
        with col3:
            atributo = st.selectbox("Atributo", options=atributos_filtrados)
        col4, col5 = st.columns(2)
        with col4:
            aql = st.number_input("Proporción esperada de unidades fuera de especificación (p)", min_value=0.0, max_value=1.0, value=1.0, step=0.01)
        with col5:
            lote = st.number_input("Tamaño del lote", min_value=0.0, max_value=1000000.0, value=1.0, step=1.0)
        col6, col7 = st.columns(2)
        with col6:
            criticidad = st.select_slider("Nivel de Criticidad", options=["Bajo", "Moderado", "Alto"])
        with col7:
            margen_error = st.select_slider("Margen de error %", options=["1.0%", "5.0%"])
    st.markdown("""
        <div style='margin-top: 10px; font-size: 13px; color: #666; text-align: justify;'>
            <p><sup style='font-size: 10px; vertical-align: super;'>1</sup> Si no tienes acceso a datos históricos del valor de <i>p</i>, consulta el MIA y establece <i>p = AQL</i> (como número) para el atributo analizado. Si sabes que el proceso se encuentra bajo control para este atributo, puedes sugerir que <i>p = AQL / 2</i> o <i>p = AQL / 3</i>.</p>
            <p><sup style='font-size: 10px; vertical-align: super;'>2</sup> El margen de error (<i>E</i>) representa cuánta precisión deseas tener al estimar la proporción de incumplimientos (<i>p</i>). En otras palabras, define la precisión estadística del muestreo. Si tu proceso es muy variable para este atributo, o el atributo es un CQA, selecciona 1%. De lo contrario, selecciona 5%.</p>
        </div>
    """, unsafe_allow_html=True)


with centro[1]:
    if st.button('⚠️ Advertencia'):
        st.warning('Por favor, revise cuidadosamente el tamaño del lote antes de continuar.')
